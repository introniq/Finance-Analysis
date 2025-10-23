import base64
import io
from collections import defaultdict
import logging
import warnings
import json
import numpy as np
import pandas as pd
from flask import Flask, request, jsonify, Response, send_from_directory
from flask_cors import CORS
import dash
from dash import dcc, html, dash_table, Input, Output, State
import dash_bootstrap_components as dbc
import yfinance as yf
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from threading import Thread, Lock
import time
import re
from openpyxl import load_workbook
from datetime import datetime, timedelta
import traceback
import queue
from flask.json.provider import DefaultJSONProvider
from pandas._libs.tslibs import NaTType

class SafeJSONProvider(DefaultJSONProvider):
    def default(self, obj):
        if isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        if isinstance(obj, NaTType):
            return None
        if isinstance(obj, np.datetime64):
            return None if pd.isna(obj) else pd.Timestamp(obj).isoformat()
        return super().default(obj)

def clean_for_json(obj):
    if isinstance(obj, (np.integer, np.floating)):
        return None if (np.isnan(obj) or np.isinf(obj)) else obj.item()
    if isinstance(obj, (np.datetime64, pd.Timestamp)):
        return None if pd.isna(obj) else obj.isoformat()
    if obj is None or (isinstance(obj, float) and (np.isnan(obj) or np.isinf(obj))):
        return None
    if isinstance(obj, dict):
        return {str(k): clean_for_json(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [clean_for_json(i) for i in obj]
    if isinstance(obj, pd.DataFrame):
        obj = obj.replace({np.nan: None, pd.NaT: None})
        return clean_for_json(obj.to_dict('records'))
    if isinstance(obj, pd.Series):
        return clean_for_json(obj.replace({np.nan: None, pd.NaT: None}).tolist())
    if isinstance(obj, np.ndarray):
        return [clean_for_json(i) for i in obj.tolist()]
    return obj

warnings.filterwarnings('ignore', category=FutureWarning)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
warnings.filterwarnings('ignore', message="Could not infer format")
warnings.filterwarnings('ignore', message="DataFrame columns are not unique")

dash_app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP], suppress_callback_exceptions=True)
server = Flask(__name__)
server.json = SafeJSONProvider(server)
CORS(server)
dash_app.server = server

symbol_queues = {}
queue_lock = Lock()

def streamer_worker(symbol, q):
    from datetime import datetime
    from zoneinfo import ZoneInfo
    logging.info("[STREAM] Worker started for %s", symbol)
    def is_market_open():
        now = datetime.now(ZoneInfo("Asia/Kolkata"))
        market_open = datetime.time(9, 15)
        market_close = datetime.time(15, 30)
        return now.weekday() < 5 and market_open <= now.time() <= market_close
    while True:
        if not is_market_open():
            logging.info("[STREAM] Market closed for %s, waiting...", symbol)
            q.put({'heartbeat': 1, 'message': 'Market closed', 'symbol': symbol}, block=False)
            time.sleep(60)
            continue
        try:
            tk = yf.Ticker(symbol)
            hist = tk.history(period='1d', interval='1m')
            if hist.empty:
                logging.warning("[STREAM] Empty history for %s", symbol)
                time.sleep(30)
                continue
            last = hist.iloc[-1]
            payload = {
                'price': float(last['Close']),
                'change': float(last['Close'] - last['Open']),
                'change_pct': float((last['Close'] / last['Open'] - 1) * 100),
                'volume': int(last['Volume']),
                'open': float(last['Open']),
                'high': float(last['High']),
                'low': float(last['Low']),
                'market_cap': tk.info.get('marketCap', 0),
                'pcr_update': 0.67,
                'timestamp': datetime.now().isoformat(),
                'symbol': symbol
            }
            q.put(payload, block=False)
            logging.info("[STREAM] Tick for %s: %s", symbol, payload['price'])
        except Exception as e:
            logging.exception("[STREAM] Tick error for %s: %s", symbol, e)
        time.sleep(30)

def get_or_create_queue(symbol):
    with queue_lock:
        if symbol not in symbol_queues:
            q = queue.Queue(maxsize=10)
            symbol_queues[symbol] = q
            Thread(target=streamer_worker, args=(symbol, q), daemon=True).start()
            logging.info("[STREAM] New queue and worker created for %s", symbol)
        return symbol_queues[symbol]

def parse_csv_robust(decoded):
    try:
        text = decoded.decode('utf-8-sig')
        lines = text.split('\n')
        header_index = -1
        sep = ','
        for i, line in enumerate(lines):
            if line.strip():
                if line.startswith('Date:') or line.startswith('Generated by'):
                    continue
                fields = [f.strip() for f in line.split(',')]
                if len(fields) >= 5:
                    first_field = fields[0].lower()
                    if 'date' in first_field:
                        header_index = i
                        sep = ','
                        break
                if header_index == -1:
                    fields = [f.strip() for f in line.split('\t')]
                    if len(fields) >= 5:
                        first_field = fields[0].lower()
                        if 'date' in first_field:
                            header_index = i
                            sep = '\t'
                            break
        if header_index == -1:
            for i, line in enumerate(lines):
                if line.strip() and not line.startswith('Date:') and not line.startswith('Generated by'):
                    header_index = i
                    sep = ',' if ',' in line else '\t'
                    break
        if header_index == -1:
            raise ValueError("No valid header line found")
        header_line = lines[header_index].strip()
        header = [h.strip() for h in header_line.split(sep) if h.strip()]
        data_lines = []
        for line in lines[header_index + 1:]:
            stripped_line = line.strip()
            if stripped_line and not stripped_line.startswith('Date:') and not stripped_line.startswith('Generated by'):
                data_lines.append(stripped_line)
        data = []
        for line in data_lines:
            row = [f.strip() for f in line.split(sep)]
            if len(row) >= len(header):
                data.append(row[:len(header)])
            elif len(row) > 0:
                data.append(row + [''] * (len(header) - len(row)))
        if not data:
            raise ValueError("No data rows found")
        df = pd.DataFrame(data, columns=header)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        return df
    except Exception as e:
        logging.error(f"Error parsing CSV: {e}")
        return pd.DataFrame()

def parse_excel_robust(decoded):
    try:
        wb = load_workbook(io.BytesIO(decoded), data_only=True)
        ws = wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data.append(row)
        df = pd.DataFrame(data)
        if df.empty:
            raise ValueError("Empty Excel sheet")
        header_row = None
        for i in range(len(df)):
            if pd.isna(df.iloc[i, 0]):
                continue
            cell_val = str(df.iloc[i, 0]).strip().lower()
            if 'date' in cell_val or re.match(r'^\d{5,}$', str(df.iloc[i, 0])):
                header_row = i
                break
        if header_row is None:
            header_row = 0
        df.columns = [str(c).strip() if c is not None else f'Col_{j}' for j, c in enumerate(df.iloc[header_row])]
        df = df.iloc[header_row + 1:].reset_index(drop=True)
        df = df.dropna(how='all')
        if 'Date' in df.columns:
            def convert_excel_date(val):
                try:
                    if isinstance(val, (int, float)) and val > 1:
                        return pd.to_datetime(val, unit='D', origin='1899-12-30')
                    return pd.to_datetime(val, errors='coerce')
                except:
                    return pd.NaT
            df['Date'] = df['Date'].apply(convert_excel_date)
        return df
    except Exception as e:
        logging.error(f"Error parsing Excel: {e}")
        return pd.DataFrame()

def parse_contents(contents, filename):
    content_type, content_string = contents.split(',')
    decoded = base64.b64decode(content_string)
    try:
        if filename.lower().endswith('.csv'):
            df = parse_csv_robust(decoded)
        elif filename.lower().endswith(('.xlsx', '.xls')):
            df = parse_excel_robust(decoded)
        else:
            raise ValueError("Unsupported file type")
        if df.empty:
            return df
        df.columns = [col.strip() for col in df.columns]
        seen = set()
        unique_cols = []
        for col in df.columns:
            if col not in seen:
                seen.add(col)
                unique_cols.append(col)
        df = df[unique_cols]
        date_cols = [col for col in df.columns if 'date' in col.lower()]
        if date_cols:
            date_col = date_cols[0]
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce', dayfirst=True)
            df = df.dropna(subset=[date_col]).sort_values(date_col).reset_index(drop=True)
            df = df.rename(columns={date_col: 'Date'})
        numeric_columns = [
            'Open', 'High', 'Low', 'Vwap', 'Close', 'Change %', 'Cumulative Future OI',
            'Future OI Change %', 'Cumulative Call OI', 'Cumulative Put OI', 'PCR',
            'Volume', 'Avg. Volume (5 Days)', 'Volume Times(x)', 'Delivery',
            'Avg. Delivery (5 Days)', 'Delivery Times(x)', 'Amount'
        ]
        for col in numeric_columns:
            if col in df.columns:
                s = df[col].astype(str) if df[col].dtype != 'object' else df[col]
                df[col] = (s.str.replace('%', '', regex=False)
                           .str.replace(',', '', regex=False)
                           .str.replace('₹', '', regex=False)
                           .str.strip())
                df[col] = pd.to_numeric(df[col], errors='coerce')
        return df
    except Exception as e:
        logging.error(f"Error parsing contents: {e}")
        return pd.DataFrame()

def extract_symbol(df):
    if df.empty:
        return "UNKNOWN"
    symbol = "UNKNOWN"
    symbol_cols = ['Symbol', 'Stock Name', 'Ticker', 'Exch']
    for col in symbol_cols:
        if col in df.columns:
            try:
                val = str(df[col].dropna().iloc[0] if not df[col].dropna().empty else "").upper().strip()
                if val:
                    cleaned = re.sub(r'[^A-Z0-9]', '', val.split()[0])
                    if cleaned:
                        symbol = cleaned
                        break
            except Exception as e:
                logging.warning(f"Error extracting symbol from column {col}: {e}")
                continue
    symbol_map = {
        'INDUS': 'INDUSTOWER',
        'HINDUNILVR': 'HINDUNILVR',
        'DABUR': 'DABUR',
        'LTF': 'LTF',
        'APLAPLO': 'APLAPOLLO'
    }
    return symbol_map.get(symbol, symbol)

def fetch_yfinance_data(symbol):
    try:
        if not symbol or symbol == 'UNKNOWN':
            logging.warning("Empty or unknown symbol provided to yfinance")
            return {'market_cap_rs': 0, 'market_cap_cr': 0, 'outstanding_shares': None, 'current_price': 'N/A', '52w_high': 'N/A', '52w_low': 'N/A', 'historical_data': pd.DataFrame()}
        sym_clean = re.sub(r'[^A-Z0-9]', '', str(symbol).strip().upper())
        if len(sym_clean) < 2:
            raise ValueError('Invalid symbol after cleaning')
        candidate = f"{sym_clean}.NS"
        ticker = yf.Ticker(candidate)
        info = ticker.info
        hist = ticker.history(period='1y')
        if hist.empty:
            raise ValueError("No historical data found")
        market_cap_rs = info.get('marketCap', 0)
        market_cap_cr = market_cap_rs / 10000000 if market_cap_rs else 0
        outstanding_shares = info.get('sharesOutstanding')
        current_price = info.get('currentPrice', hist['Close'].iloc[-1] if not hist.empty else 'N/A')
        _52w_high = info.get('fiftyTwoWeekHigh', hist['High'].max() if not hist.empty else 'N/A')
        _52w_low = info.get('fiftyTwoWeekLow', hist['Low'].min() if not hist.empty else 'N/A')
        return {
            'market_cap_rs': market_cap_rs,
            'market_cap_cr': market_cap_cr,
            'outstanding_shares': outstanding_shares,
            'current_price': current_price,
            '52w_high': _52w_high,
            '52w_low': _52w_low,
            'historical_data': hist
        }
    except Exception as e:
        logging.error(f"Error fetching yfinance data for {symbol}: {e}")
        return {'market_cap_rs': 0, 'market_cap_cr': 0, 'outstanding_shares': None, 'current_price': 'N/A', '52w_high': 'N/A', '52w_low': 'N/A', 'historical_data': pd.DataFrame()}

class StockAnalyzer:
    def __init__(self, df, market_cap_cr, outstanding_shares, yf_data=None, symbol=None):
        self.df = df if not df.empty else pd.DataFrame()
        self.market_cap_cr = market_cap_cr or 0
        self.outstanding_shares = outstanding_shares
        self.yf_data = yf_data or {}
        self.symbol = symbol or "UNKNOWN"
        self.patterns = []
        self.qualifying = pd.DataFrame()
        self.labels = None
        self.patterns_array = None
        self.scaler = None
        self.model = None

    def compute_indicators(self, rsi_period=14, adx_period=14):
        if self.df.empty or 'Close' not in self.df.columns:
            logging.warning("No 'Close' column or empty df, skipping indicators")
            return
        close = self.df['Close'].ffill()
        high = self.df.get('High', close)
        low = self.df.get('Low', close)
        volume = self.df.get('Volume', pd.Series(1, index=self.df.index))
        self.df['MA5'] = close.rolling(5).mean()
        delta = close.diff()
        gain = delta.where(delta > 0, 0).rolling(window=rsi_period).mean()
        loss = -delta.where(delta < 0, 0).rolling(window=rsi_period).mean()
        rs = gain / loss
        self.df['RSI'] = 100 - (100 / (1 + rs))
        ema12 = close.ewm(span=12).mean()
        ema26 = close.ewm(span=26).mean()
        self.df['MACD'] = ema12 - ema26
        self.df['MACD_Signal'] = self.df['MACD'].ewm(span=9).mean()
        self.df['BB_Mid'] = close.rolling(20).mean()
        bb_std = close.rolling(20).std()
        self.df['BB_Upper'] = self.df['BB_Mid'] + (bb_std * 2)
        self.df['BB_Lower'] = self.df['BB_Mid'] - (bb_std * 2)
        low_min = low.rolling(rsi_period).min()
        high_max = high.rolling(rsi_period).max()
        self.df['Stoch_K'] = 100 * ((close - low_min) / (high_max - low_min))
        self.df['Stoch_D'] = self.df['Stoch_K'].rolling(3).mean()
        up = high - high.shift(1)
        down = low.shift(1) - low
        plus_dm = pd.Series(np.where((up > down) & (up > 0), up, 0), index=up.index)
        minus_dm = pd.Series(np.where((down > up) & (down > 0), down, 0), index=down.index)
        tr = pd.concat([
            high - low,
            abs(high - close.shift(1)),
            abs(low - close.shift(1))
        ], axis=1).max(axis=1)
        plus_di = 100 * (plus_dm.rolling(adx_period).mean() / tr.rolling(adx_period).mean())
        minus_di = 100 * (minus_dm.rolling(adx_period).mean() / tr.rolling(adx_period).mean())
        dx = (abs(plus_di - minus_di) / (plus_di + minus_di)) * 100
        self.df['ADX'] = dx.rolling(adx_period).mean()
        self.df['+DI'] = plus_di
        self.df['-DI'] = minus_di
        typical_price = (high + low + close) / 3
        self.df['VWAP'] = (typical_price * volume).cumsum() / volume.cumsum()
        if 'Delivery' in self.df.columns:
            self.df['Delivery_Pct'] = (self.df['Delivery'] / self.df['Volume']) * 100
            self.df['Delivery_Value'] = self.df['Delivery'] * close
            self.df['Sum_Delivery_Value'] = self.df['Delivery_Value'].cumsum()
        else:
            self.df['Delivery'] = volume * 0.5
            self.df['Delivery_Pct'] = 50.0
            self.df['Delivery_Value'] = self.df['Delivery'] * close
            self.df['Sum_Delivery_Value'] = self.df['Delivery_Value'].cumsum()
        if 'Cumulative Future OI' in self.df.columns:
            self.df['OI_Change_Pct'] = self.df['Cumulative Future OI'].pct_change() * 100
        else:
            self.df['Cumulative Future OI'] = volume * 0.1
            self.df['OI_Change_Pct'] = self.df['Cumulative Future OI'].pct_change() * 100
        self.df['Wyckoff_Event'] = 'Neutral'
        bull_conditions = (close > self.df['MA5']) & (volume > volume.rolling(20).mean())
        bear_conditions = (close < self.df['MA5']) & (volume > volume.rolling(20).mean())
        self.df.loc[bull_conditions, 'Wyckoff_Event'] = 'Sign of Strength'
        self.df.loc[bear_conditions, 'Wyckoff_Event'] = 'Sign of Weakness'

    def _edge_diagonal_oi(self, oi_df, current_price, edge_pct=15):
        if oi_df.empty or 'Price Level' not in oi_df.columns or 'OI' not in oi_df.columns:
            return {'edge_oi': 0, 'edge_pct': 0.0, 'diagonal_oi': 0, 'diagonal_pct': 0.0}
        oi_df = oi_df.copy()
        oi_df['Price Level'] = pd.to_numeric(oi_df['Price Level'], errors='coerce')
        oi_df['OI'] = pd.to_numeric(oi_df['OI'], errors='coerce')
        total_oi = oi_df['OI'].sum()
        if total_oi == 0:
            return {'edge_oi': 0, 'edge_pct': 0.0, 'diagonal_oi': 0, 'diagonal_pct': 0.0}
        ref = float(current_price)
        dist = abs(oi_df['Price Level'] - ref) / ref * 100
        edge_mask = dist >= edge_pct
        edge_oi = oi_df.loc[edge_mask, 'OI'].sum()
        hi_thr = oi_df['Price Level'].quantile(0.80)
        lo_thr = oi_df['Price Level'].quantile(0.20)
        diagonal_mask = (oi_df['Price Level'] >= hi_thr) | (oi_df['Price Level'] <= lo_thr)
        diagonal_oi = oi_df.loc[diagonal_mask, 'OI'].sum()
        return {
            'edge_oi': int(edge_oi),
            'edge_pct': round(float(edge_oi / total_oi * 100), 2),
            'diagonal_oi': int(diagonal_oi),
            'diagonal_pct': round(float(diagonal_oi / total_oi * 100), 2)
        }

    def _cumulative_oi_open_close(self, oi_df):
        if oi_df.empty or 'Price Level' not in oi_df.columns or 'OI' not in oi_df.columns:
            return pd.DataFrame(), pd.DataFrame()
        df = oi_df.copy()
        df['Price Level'] = pd.to_numeric(df['Price Level'], errors='coerce')
        df['OI'] = pd.to_numeric(df['OI'], errors='coerce')
        df = df.sort_values('Price Level').reset_index(drop=True)
        df['Cum_OI_Up'] = df['OI'].cumsum()
        df['Cum_OI_Down'] = df['OI'][::-1].cumsum()[::-1].values
        return df[['Price Level', 'Cum_OI_Up', 'Cum_OI_Down']]

    def _oi_turnaround_point(self, oi_df, current_price):
        if oi_df.empty or 'Price Level' not in oi_df.columns or 'OI' not in oi_df.columns:
            return {'turnaround_price': 'N/A', 'turnaround_pct': 0.0}
        df = oi_df.copy()
        df['Price Level'] = pd.to_numeric(df['Price Level'], errors='coerce')
        df['OI'] = pd.to_numeric(df['OI'], errors='coerce')
        df = df.sort_values('Price Level').reset_index(drop=True)
        total = df['OI'].sum()
        if total == 0:
            return {'turnaround_price': 'N/A', 'turnaround_pct': 0.0}
        df['Cum_Pct'] = df['OI'].cumsum() / total * 100
        idx = df[df['Cum_Pct'] >= 50].index.min()
        if pd.isna(idx):
            return {'turnaround_price': 'N/A', 'turnaround_pct': 0.0}
        turnaround_price = float(df.loc[idx, 'Price Level'])
        turnaround_pct = float(df.loc[idx, 'Cum_Pct'])
        return {
            'turnaround_price': round(turnaround_price, 2),
            'turnaround_pct': round(turnaround_pct, 2)
        }

    def find_qualifying_windows(self, window_size=10):
        if len(self.df) < window_size or self.df.empty:
            self.qualifying = pd.DataFrame()
            return
        if 'Delivery' not in self.df.columns or 'Close' not in self.df.columns or 'Cumulative Future OI' not in self.df.columns:
            self.qualifying = pd.DataFrame()
            return
        self.df['Delivery_Value'] = self.df['Delivery'] * self.df['Close']
        self.df['Rolling_Delivery_Value_Sum'] = self.df['Delivery_Value'].rolling(window=window_size, min_periods=window_size).sum()
        self.df['Rolling_OI_Cum_Increase_Pct'] = (
            (self.df['Cumulative Future OI'] - self.df['Cumulative Future OI'].shift(window_size - 1)) / 
            self.df['Cumulative Future OI'].shift(window_size - 1) * 100
        ).fillna(0)
        delivery_threshold = 0.015 * self.market_cap_cr
        qualifying_mask = (
            (self.df['Rolling_Delivery_Value_Sum'] > delivery_threshold) &
            (self.df['Rolling_OI_Cum_Increase_Pct'] > 10)
        )
        self.qualifying = self.df[qualifying_mask].copy()
        if not self.qualifying.empty:
            self.qualifying['Window_Start'] = self.qualifying['Date'].shift(window_size - 1)
            self.qualifying['Delivery_Sum_Cr'] = (self.qualifying['Rolling_Delivery_Value_Sum'] / 10000000).round(2)
            self.qualifying['Delivery_vs_MC_Pct'] = (self.qualifying['Rolling_Delivery_Value_Sum'] / (self.market_cap_cr * 10000000) * 100).round(2)
            self.qualifying['OI_Increase_Pct'] = self.qualifying['Rolling_OI_Cum_Increase_Pct'].round(2)
            self.qualifying = self.qualifying[['Window_Start', 'Date', 'Close', 'Delivery_Sum_Cr', 'Delivery_vs_MC_Pct', 'OI_Increase_Pct', 'Wyckoff_Event']]

    def cluster_patterns(self, algorithm='kmeans', n_clusters=3):
        window_size = 10
        if len(self.df) < window_size or self.df.empty:
            self.labels = np.array([])
            return
        self.patterns = []
        for i in range(window_size, len(self.df)):
            window = self.df.iloc[i - window_size:i]
            if len(window) < window_size:
                continue
            ret_mean = window['Close'].pct_change().mean() if 'Close' in window.columns else 0
            vol_mean = window['Volume'].pct_change().mean() if 'Volume' in window.columns else 0
            oi_mean = window['OI_Change_Pct'].mean() if 'OI_Change_Pct' in window.columns else 0
            pattern = [ret_mean, vol_mean, oi_mean]
            self.patterns.append(pattern)
        self.patterns_array = np.array(self.patterns)
        if len(self.patterns_array) == 0:
            self.labels = np.array([])
            return
        if np.isnan(self.patterns_array).any():
            self.patterns_array = np.nan_to_num(self.patterns_array)
        self.scaler = StandardScaler()
        scaled = self.scaler.fit_transform(self.patterns_array)
        if algorithm == 'kmeans':
            self.model = KMeans(n_clusters=n_clusters, random_state=42)
            self.labels = self.model.fit_predict(scaled)
        else:
            self.labels = np.array([])

    def plot_clusters(self, symbol, algo):
        if (self.patterns_array is None or len(self.patterns_array) == 0 or
            len(self.labels) == 0 or len(self.patterns_array) != len(self.labels)):
            fig = go.Figure()
            fig.update_layout(title=f'No data for clusters {symbol}')
            return fig
        fig = px.scatter(
            x=self.patterns_array[:, 0], y=self.patterns_array[:, 1],
            color=self.labels.astype(str),
            labels={'x': 'Avg Return', 'y': 'Avg Vol Change', 'color': 'Cluster'}
        )
        fig.update_layout(title=f'Pattern Clusters for {symbol} ({algo.upper()})')
        return fig

    def analyze_pcr_trends(self, symbol):
        if 'PCR' not in self.df.columns or self.df.empty:
            fig = go.Figure()
            fig.update_layout(title=f'No PCR data for {symbol}')
            return fig, {'mean': 'N/A', 'latest': 'N/A', 'trend': 'N/A'}
        fig = px.line(self.df, x='Date', y='PCR', title=f'PCR Trends for {symbol}')
        stats = {
            'mean': round(self.df['PCR'].mean(), 2),
            'latest': round(self.df['PCR'].iloc[-1], 2),
            'trend': 'Rising' if self.df['PCR'].iloc[-1] > self.df['PCR'].iloc[-5] else 'Falling'
        }
        return fig, stats

    def analyze_wyckoff(self, symbol):
        overview = "Wyckoff analysis detects accumulation/distribution phases based on price-volume action."
        recent = self.df['Wyckoff_Event'].iloc[-1] if not self.df.empty and 'Wyckoff_Event' in self.df.columns else 'Neutral'
        wyckoff_data = {'overview': overview, 'recent': recent}
        fig = px.line(self.df, x='Date', y='Close', title=f'Wyckoff Analysis for {symbol}')
        if not self.df.empty and 'Wyckoff_Event' in self.df.columns:
            last_date = self.df['Date'].iloc[-1]
            last_close = self.df['Close'].iloc[-1]
            fig.add_annotation(x=last_date, y=last_close, text=recent, showarrow=True)
        return fig, wyckoff_data

    def _build_historical_supply_map(self, df, price_bins):
        hist_supply = defaultdict(int)
        if df.empty or 'Close' not in df.columns or 'Volume' not in df.columns:
            return hist_supply
        df = df.copy()
        close_fill = df['Close'].ffill()
        df['Price_Bin'] = pd.cut(close_fill, bins=price_bins)
        for idx, row in df.iterrows():
            if pd.isna(row['Price_Bin']):
                continue
            mid = row['Price_Bin'].mid
            vol = row['Volume']
            if vol > df['Volume'].quantile(0.50):
                hist_supply[mid] += 1
        return hist_supply

    def _attach_supply_flag_per_row(self, sub_df):
        price_min = sub_df['Low'].min() if 'Low' in sub_df.columns else sub_df['Close'].min()
        price_max = sub_df['High'].max() if 'High' in sub_df.columns else sub_df['Close'].max()
        if price_min == price_max:
            price_bins = np.array([price_min, price_min + 1e-6])
        else:
            price_bins = np.linspace(price_min, price_max, 20)
        close_fill = sub_df['Close'].ffill()
        sub_df['Price_Bin'] = pd.cut(close_fill, bins=price_bins)
        hist_map = self._build_historical_supply_map(sub_df, price_bins)
        sub_df['Supply_Check'] = sub_df.apply(
            lambda r: 'Heavy Supply' if hist_map.get(r['Price_Bin'].mid, 0) > 0 else 'Balanced', axis=1
        )
        sub_df['Historical_Supply_Count'] = sub_df['Price_Bin'].map(
            lambda b: hist_map.get(b.mid, 0)
        )
        return sub_df

    def compute_volume_profile(self, date_ranges, peak_diff_dates=None):
        if ('Close' not in self.df.columns or 'Volume' not in self.df.columns
                or self.df.empty or not date_ranges):
            fig = go.Figure()
            empty_df = pd.DataFrame(
                columns=['Date', 'Price Level', 'Volume', 'Percentage (%)',
                        'Supply_Check', 'Historical_Supply_Count'])
            return (fig, {}, [], empty_df, 'N/A', 0, empty_df,
                    {'high_vol_pct': 0.0, 'low_vol_pct': 100.0,
                    'imbalance': 0.0, 'supply_check': 'No Data'},
                    'N/A', 'N/A', 0, 0)
        fig = go.Figure()
        vp_dfs, pocs, total_volumes = {}, {}, {}
        supply_checks = {}
        va_highs, va_lows, va_diffs = {}, {}, {}
        va_vol_pcts, cum_dfs = {}, {}
        for start, end in date_ranges:
            if pd.isna(start) or pd.isna(end):
                continue
            mask = (self.df['Date'] >= start) & (self.df['Date'] <= end)
            sub_df = self.df[mask].copy()
            if sub_df.empty:
                continue
            price_min = sub_df['Low'].min() if 'Low' in sub_df.columns else sub_df['Close'].min()
            price_max = sub_df['High'].max() if 'High' in sub_df.columns else sub_df['Close'].max()
            if price_min == price_max:
                price_bins = np.array([price_min, price_min + 1e-6])
            else:
                price_bins = np.linspace(price_min, price_max, 20)
            close_fill = sub_df['Close'].ffill()
            sub_df['Price_Bin'] = pd.cut(close_fill, bins=price_bins)
            sub_df = sub_df.dropna(subset=['Price_Bin'])
            if sub_df.empty:
                continue
            vol_profile = sub_df.groupby('Price_Bin')['Volume'].sum()
            mids = [interval.mid for interval in vol_profile.index]
            hist_map = self._build_historical_supply_map(sub_df, price_bins)
            price_df = pd.DataFrame({
                'Price_Bin': [str(iv) for iv in vol_profile.index],
                'Price Level': mids,
                'Volume': vol_profile.values
            }).sort_values('Price Level').reset_index(drop=True)
            if len(price_df) == 0:
                continue
            total_vol = price_df['Volume'].sum()
            target_vol = total_vol * 0.7
            poc_idx = price_df['Volume'].idxmax()
            poc_price = price_df.loc[poc_idx, 'Price Level'] if pd.notna(poc_idx) else 'N/A'
            current_vol = price_df.loc[poc_idx, 'Volume'] if pd.notna(poc_idx) else 0
            va_start = poc_idx if pd.notna(poc_idx) else 0
            va_end = poc_idx if pd.notna(poc_idx) else 0
            while current_vol < target_vol and (va_start > 0 or va_end < len(price_df) - 1):
                above_idx = va_end + 1
                below_idx = va_start - 1
                above_vol = price_df.loc[above_idx, 'Volume'] if above_idx < len(price_df) else 0
                below_vol = price_df.loc[below_idx, 'Volume'] if below_idx >= 0 else 0
                if above_vol >= below_vol and above_idx < len(price_df):
                    va_end = above_idx
                    current_vol += above_vol
                elif below_idx >= 0:
                    va_start = below_idx
                    current_vol += below_vol
                else:
                    break
            va_low = price_df.loc[va_start, 'Price Level'] if pd.notna(poc_idx) else 'N/A'
            va_high = price_df.loc[va_end, 'Price Level'] if pd.notna(poc_idx) else 'N/A'
            va_diff = (va_high - va_low) if pd.notna(poc_idx) else 0
            if peak_diff_dates and len(peak_diff_dates) == 2:
                try:
                    user_start = pd.to_datetime(peak_diff_dates[0])
                    user_end = pd.to_datetime(peak_diff_dates[1])
                    user_mask = (sub_df['Date'] >= user_start) & (sub_df['Date'] <= user_end)
                    user_slice = sub_df[user_mask]
                    if not user_slice.empty:
                        peak_high = user_slice['High'].max() if 'High' in user_slice.columns else user_slice['Close'].max()
                        peak_low = user_slice['Low'].min() if 'Low' in user_slice.columns else user_slice['Close'].min()
                        va_diff = float(peak_high - peak_low)
                        peak_diff = va_diff
                    else:
                        peak_diff = va_diff
                except Exception:
                    peak_diff = va_diff
            else:
                peak_diff = va_diff
            va_vol = price_df.loc[va_start:va_end, 'Volume'].sum() if pd.notna(poc_idx) else 0
            va_vol_pct = (va_vol / total_vol * 100) if total_vol > 0 else 0
            sub_df = self._attach_supply_flag_per_row(sub_df)
            vp_df = (sub_df[['Date', 'Close', 'Volume', 'Supply_Check', 'Historical_Supply_Count']]
                    .rename(columns={'Close': 'Price Level'})
                    .assign(**{'Percentage (%)': lambda d: (d['Volume'] / total_vol * 100).round(2)})
                    .sort_values('Volume', ascending=False)
                    .reset_index(drop=True))
            vp_df['Date'] = vp_df['Date'].dt.strftime('%Y-%m-%d')
            vp_df['Cum_Volume'] = vp_df['Volume'].cumsum()
            vp_df['Cum_Pct'] = (vp_df['Cum_Volume'] / total_vol * 100).round(2) if total_vol > 0 else 0
            price_df['Cum_Vol_Bottom_Up'] = price_df['Volume'].cumsum()
            price_df['Cum_Pct_Bottom_Up'] = (price_df['Cum_Vol_Bottom_Up'] / total_vol * 100).round(2) if total_vol > 0 else 0
            high_to_low = price_df.sort_values('Price Level', ascending=False)
            high_to_low['Cum_Vol_Top_Down'] = high_to_low['Volume'].cumsum()
            high_to_low['Cum_Pct_Top_Down'] = (high_to_low['Cum_Vol_Top_Down'] / total_vol * 100).round(2) if total_vol > 0 else 0
            cum_df = price_df.merge(
                high_to_low[['Price Level', 'Cum_Vol_Top_Down', 'Cum_Pct_Top_Down']],
                on='Price Level', suffixes=('_bottom', '_top'))
            prices = price_df['Price Level'].tolist()
            if len(prices) == 0:
                high_vol_pct, low_vol_pct, imbalance, supply_check_status = 0.0, 100.0, 0.0, 'No Data'
            else:
                price_threshold_high = np.percentile(prices, 70)
                high_vol = sum(vol for p, vol in zip(prices, price_df['Volume'].tolist()) if p >= price_threshold_high)
                low_vol = total_vol - high_vol
                high_vol_pct = (high_vol / total_vol * 100) if total_vol > 0 else 0
                low_vol_pct = 100 - high_vol_pct
                supply_check_status = "Heavy Supply" if high_vol_pct > 35 else "Balanced" if high_vol_pct > 25 else "Demand Heavy"
                imbalance = high_vol_pct - low_vol_pct
            key = f"{start.date()} to {end.date()}"
            vp_dfs[key] = vp_df
            pocs[key] = poc_price
            total_volumes[key] = total_vol
            va_highs[key] = va_high
            va_lows[key] = va_low
            va_diffs[key] = va_diff
            va_vol_pcts[key] = va_vol_pct
            cum_dfs[key] = cum_df
            supply_checks[key] = {
                'high_vol_pct': round(high_vol_pct, 1),
                'low_vol_pct': round(low_vol_pct, 1),
                'imbalance': round(imbalance, 1),
                'supply_check': supply_check_status
            }
            fig.add_trace(go.Bar(x=mids, y=vol_profile.values, name=key))
        fig.update_layout(
            title='Volume Profile: Price Levels vs. Total Volume',
            xaxis_title='Price Level',
            yaxis_title='Volume',
            bargap=0.1,
            showlegend=True,
            height=500
        )
        main_key = list(vp_dfs.keys())[0] if vp_dfs else None
        if main_key:
            return (fig, {}, [], vp_dfs[main_key], pocs[main_key],
                    total_volumes[main_key], cum_dfs[main_key], supply_checks[main_key],
                    va_highs[main_key], va_lows[main_key], va_diffs[main_key], va_vol_pcts[main_key], peak_diff)
        else:
            empty_df = pd.DataFrame(
                columns=['Date', 'Price Level', 'Volume', 'Percentage (%)',
                        'Supply_Check', 'Historical_Supply_Count'])
            return (fig, {}, [], empty_df, 'N/A', 0, empty_df,
                    {'high_vol_pct': 0.0, 'low_vol_pct': 100.0,
                    'imbalance': 0.0, 'supply_check': 'No Data'},
                    'N/A', 'N/A', 0, 0)

    def compute_oi_profile(self, date_ranges):
        oi_cols = ['Cumulative Future OI', 'Cumulative Call OI', 'Cumulative Put OI']
        oi_col = None
        for col in oi_cols:
            if col in self.df.columns:
                oi_col = col
                break
        if (oi_col is None or 'Close' not in self.df.columns or
            self.df.empty or not date_ranges):
            fig = go.Figure()
            empty_df = pd.DataFrame(columns=['Price Level', 'OI', 'Percentage (%)'])
            return (fig, {}, [], empty_df, 'N/A', 0, empty_df,
                    {'high_oi_pct': 0.0, 'low_oi_pct': 100.0, 'imbalance': 0.0, 'supply_check': 'No Data'},
                    'N/A', 'N/A', 0, 0)
        fig = go.Figure()
        op_dfs = {}
        pocs = {}
        total_ois = {}
        supply_checks = {}
        va_highs = {}
        va_lows = {}
        va_diffs = {}
        va_oi_pcts = {}
        cum_dfs = {}
        for start, end in date_ranges:
            if pd.isna(start) or pd.isna(end):
                continue
            mask = (self.df['Date'] >= start) & (self.df['Date'] <= end)
            sub_df = self.df[mask].copy()
            if sub_df.empty:
                continue
            price_min = sub_df['Low'].min() if 'Low' in sub_df.columns else sub_df['Close'].min()
            price_max = sub_df['High'].max() if 'High' in sub_df.columns else sub_df['Close'].max()
            if price_min == price_max:
                price_bins = np.array([price_min, price_min + 1e-6])
            else:
                price_bins = np.linspace(price_min, price_max, 20)
            close_fill = sub_df['Close'].ffill()
            sub_df['Price_Bin'] = pd.cut(close_fill, bins=price_bins)
            sub_df = sub_df.dropna(subset=['Price_Bin'])
            if sub_df.empty:
                key = f"{start.date()} to {end.date()}"
                op_dfs[key] = pd.DataFrame(columns=['Price Level', 'OI', 'Percentage (%)'])
                pocs[key] = 'N/A'
                total_ois[key] = 0
                supply_checks[key] = {'high_oi_pct': 0.0, 'low_oi_pct': 100.0, 'imbalance': 0.0, 'supply_check': 'No Data'}
                va_highs[key] = 'N/A'
                va_lows[key] = 'N/A'
                va_diffs[key] = 0
                va_oi_pcts[key] = 0
                cum_dfs[key] = pd.DataFrame()
                continue
            oi_profile = sub_df.groupby('Price_Bin')[oi_col].sum()
            mids = [interval.mid for interval in oi_profile.index]
            price_df = pd.DataFrame({
                'Price_Bin': [str(interval) for interval in oi_profile.index],
                'Price Level': mids,
                'OI': oi_profile.values
            }).sort_values('Price Level').reset_index(drop=True)
            if len(price_df) == 0:
                key = f"{start.date()} to {end.date()}"
                op_dfs[key] = pd.DataFrame(columns=['Price Level', 'OI', 'Percentage (%)'])
                pocs[key] = 'N/A'
                total_ois[key] = 0
                supply_checks[key] = {'high_oi_pct': 0.0, 'low_oi_pct': 100.0, 'imbalance': 0.0, 'supply_check': 'No Data'}
                va_highs[key] = 'N/A'
                va_lows[key] = 'N/A'
                va_diffs[key] = 0
                va_oi_pcts[key] = 0
                cum_dfs[key] = pd.DataFrame()
                continue
            total_oi = price_df['OI'].sum()
            target_oi = total_oi * 0.7
            poc_idx = price_df['OI'].idxmax()
            if pd.isna(poc_idx):
                poc_price = 'N/A'
                poc_oi = 0
            else:
                poc_price = price_df.loc[poc_idx, 'Price Level']
                poc_oi = price_df.loc[poc_idx, 'OI']
            current_oi = poc_oi
            if pd.isna(poc_idx):
                va_start = len(price_df)
                va_end = -1
            else:
                va_start = poc_idx
                va_end = poc_idx
            while current_oi < target_oi and (va_start > 0 or va_end < len(price_df) - 1):
                above_idx = va_end + 1
                below_idx = va_start - 1
                above_oi = price_df.loc[above_idx, 'OI'] if above_idx < len(price_df) else 0
                below_oi = price_df.loc[below_idx, 'OI'] if below_idx >= 0 else 0
                if above_oi >= below_oi and above_idx < len(price_df):
                    va_end = above_idx
                    current_oi += above_oi
                elif below_idx >= 0:
                    va_start = below_idx
                    current_oi += below_oi
                else:
                    break
            if pd.isna(poc_idx) or total_oi == 0:
                va_low = 'N/A'
                va_high = 'N/A'
                va_diff = 0
                va_oi = 0
                va_oi_pct = 0
            else:
                va_low = price_df.loc[va_start, 'Price Level']
                va_high = price_df.loc[va_end, 'Price Level']
                va_diff = va_high - va_low
                va_oi = price_df.loc[va_start:va_end, 'OI'].sum()
                va_oi_pct = (va_oi / total_oi * 100) if total_oi > 0 else 0
            percentages = (oi_profile.values / total_oi * 100).round(2) if total_oi > 0 else np.zeros(len(oi_profile.values))
            op_df = pd.DataFrame({
                'Price Level': [f"{mid:.2f}" for mid in mids],
                'OI': oi_profile.values,
                'Percentage (%)': percentages
            }).sort_values('OI', ascending=False).reset_index(drop=True)
            op_df['Cum_OI'] = op_df['OI'].cumsum()
            op_df['Cum_Pct'] = (op_df['Cum_OI'] / total_oi * 100).round(2) if total_oi > 0 else 0
            price_df['Cum_OI_Bottom_Up'] = price_df['OI'].cumsum()
            price_df['Cum_Pct_Bottom_Up'] = (price_df['Cum_OI_Bottom_Up'] / total_oi * 100).round(2) if total_oi > 0 else 0
            high_to_low = price_df.sort_values('Price Level', ascending=False)
            high_to_low['Cum_OI_Top_Down'] = high_to_low['OI'].cumsum()
            high_to_low['Cum_Pct_Top_Down'] = (high_to_low['Cum_OI_Top_Down'] / total_oi * 100).round(2) if total_oi > 0 else 0
            cum_df = price_df.merge(high_to_low[['Price Level', 'Cum_OI_Top_Down', 'Cum_Pct_Top_Down']], on='Price Level', suffixes=('_bottom', '_top'))
            prices = price_df['Price Level'].tolist()
            if len(prices) == 0:
                high_oi_pct = 0.0
                low_oi_pct = 100.0
                imbalance = 0.0
                supply_check_status = "No Data"
            else:
                price_threshold_high = np.percentile(prices, 70)
                high_oi = sum(oi for p, oi in zip(prices, price_df['OI'].tolist()) if p >= price_threshold_high)
                low_oi = total_oi - high_oi
                high_oi_pct = (high_oi / total_oi * 100) if total_oi > 0 else 0
                low_oi_pct = 100 - high_oi_pct
                supply_check_status = "Heavy OI Supply" if high_oi_pct > 35 else "Balanced OI" if high_oi_pct > 25 else "Demand Heavy OI"
                imbalance = high_oi_pct - low_oi_pct
            key = f"{start.date()} to {end.date()}"
            op_dfs[key] = op_df
            pocs[key] = poc_price
            total_ois[key] = total_oi
            va_highs[key] = va_high
            va_lows[key] = va_low
            va_diffs[key] = va_diff
            va_oi_pcts[key] = va_oi_pct
            cum_dfs[key] = cum_df
            supply_checks[key] = {
                'high_oi_pct': round(high_oi_pct, 1),
                'low_oi_pct': round(low_oi_pct, 1),
                'imbalance': round(imbalance, 1),
                'supply_check': supply_check_status
            }
            prices_plot = mids
            ois = oi_profile.values
            fig.add_trace(go.Bar(x=prices_plot, y=ois, name=key))
        fig.update_layout(
            title='Open Interest Profile: Price Levels vs. Total OI',
            xaxis_title='Price Level',
            yaxis_title='Open Interest',
            bargap=0.1,
            showlegend=True,
            height=500
        )
        main_key = list(op_dfs.keys())[0] if op_dfs else None
        if main_key:
            main_op_df = op_dfs.get(main_key, pd.DataFrame())
            main_cum_df = cum_dfs.get(main_key, pd.DataFrame())
            main_poi = pocs.get(main_key, 'N/A')
            main_total_oi = total_ois.get(main_key, 0)
            main_supply = supply_checks.get(main_key, {})
            main_va_high = va_highs.get(main_key, 'N/A')
            main_va_low = va_lows.get(main_key, 'N/A')
            main_va_diff = va_diffs.get(main_key, 0)
            main_va_oi_pct = va_oi_pcts.get(main_key, 0)
            return (fig, {}, [], main_op_df, main_poi, main_total_oi,
                    main_cum_df, main_supply, main_va_high, main_va_low, main_va_diff, main_va_oi_pct)
        else:
            empty_df = pd.DataFrame(columns=['Price Level', 'OI', 'Percentage (%)'])
            return (fig, {}, [], empty_df, 'N/A', 0, empty_df,
                    {'high_oi_pct': 0.0, 'low_oi_pct': 100.0, 'imbalance': 0.0, 'supply_check': 'No Data'},
                    'N/A', 'N/A', 0, 0)

    def compute_tpo_profile(self, date_ranges, period_size=30):
        if ('High' not in self.df.columns or 'Low' not in self.df.columns or
            'Close' not in self.df.columns or self.df.empty or not date_ranges):
            fig = go.Figure()
            empty_df = pd.DataFrame(columns=['Price Level', 'TPO Count', 'Percentage (%)'])
            return (fig, {}, [], empty_df, 'N/A', 0, empty_df,
                    {'high_tpo_pct': 0.0, 'low_tpo_pct': 100.0, 'imbalance': 0.0, 'supply_check': 'No Data'},
                    'N/A', 'N/A', 0, 0)
        fig = go.Figure()
        tpo_dfs = {}
        pocs = {}
        total_tpos = {}
        supply_checks = {}
        va_highs = {}
        va_lows = {}
        va_diffs = {}
        va_tpo_pcts = {}
        cum_dfs = {}
        for start, end in date_ranges:
            if pd.isna(start) or pd.isna(end):
                continue
            mask = (self.df['Date'] >= start) & (self.df['Date'] <= end)
            sub_df = self.df[mask].copy()
            if sub_df.empty:
                continue
            price_min = sub_df['Low'].min()
            price_max = sub_df['High'].max()
            if price_min == price_max:
                price_bins = np.array([price_min, price_min + 1e-6])
            else:
                price_bins = np.linspace(price_min, price_max, 20)
            tpo_profile = pd.Series(0, index=range(len(price_bins) - 1), dtype=int)
            for idx, row in sub_df.iterrows():
                low_bin = np.searchsorted(price_bins, row['Low'], side='right') - 1 if 'Low' in row else 0
                high_bin = np.searchsorted(price_bins, row['High'], side='right') - 1 if 'High' in row else len(tpo_profile) - 1
                for bin_idx in range(max(0, low_bin), min(len(tpo_profile), high_bin + 1)):
                    tpo_profile.iloc[bin_idx] += 1
            mids = [(price_bins[i] + price_bins[i + 1]) / 2 for i in range(len(price_bins) - 1)]
            price_df = pd.DataFrame({
                'Price_Bin': tpo_profile.index,
                'Price Level': mids,
                'TPO': tpo_profile.values
            }).sort_values('Price Level').reset_index(drop=True)
            if len(price_df) == 0:
                key = f"{start.date()} to {end.date()}"
                tpo_dfs[key] = pd.DataFrame(columns=['Price Level', 'TPO Count', 'Percentage (%)'])
                pocs[key] = 'N/A'
                total_tpos[key] = 0
                supply_checks[key] = {'high_tpo_pct': 0.0, 'low_tpo_pct': 100.0, 'imbalance': 0.0, 'supply_check': 'No Data'}
                va_highs[key] = 'N/A'
                va_lows[key] = 'N/A'
                va_diffs[key] = 0
                va_tpo_pcts[key] = 0
                cum_dfs[key] = pd.DataFrame()
                continue
            total_tpo = price_df['TPO'].sum()
            target_tpo = total_tpo * 0.7
            poc_idx = price_df['TPO'].idxmax()
            if pd.isna(poc_idx):
                poc_price = 'N/A'
                poc_tpo = 0
            else:
                poc_price = price_df.loc[poc_idx, 'Price Level']
                poc_tpo = price_df.loc[poc_idx, 'TPO']
            current_tpo = poc_tpo
            if pd.isna(poc_idx):
                va_start = len(price_df)
                va_end = -1
            else:
                va_start = poc_idx
                va_end = poc_idx
            while current_tpo < target_tpo and (va_start > 0 or va_end < len(price_df) - 1):
                above_idx = va_end + 1
                below_idx = va_start - 1
                above_tpo = price_df.loc[above_idx, 'TPO'] if above_idx < len(price_df) else 0
                below_tpo = price_df.loc[below_idx, 'TPO'] if below_idx >= 0 else 0
                if above_tpo >= below_tpo and above_idx < len(price_df):
                    va_end = above_idx
                    current_tpo += above_tpo
                elif below_idx >= 0:
                    va_start = below_idx
                    current_tpo += below_tpo
                else:
                    break
            if pd.isna(poc_idx) or total_tpo == 0:
                va_low = 'N/A'
                va_high = 'N/A'
                va_diff = 0
                va_tpo = 0
                va_tpo_pct = 0
            else:
                va_low = price_df.loc[va_start, 'Price Level']
                va_high = price_df.loc[va_end, 'Price Level']
                va_diff = va_high - va_low
                va_tpo = price_df.loc[va_start:va_end, 'TPO'].sum()
                va_tpo_pct = (va_tpo / total_tpo * 100) if total_tpo > 0 else 0
            percentages = (tpo_profile.values / total_tpo * 100).round(2) if total_tpo > 0 else np.zeros(len(tpo_profile.values))
            tpo_df = pd.DataFrame({
                'Price Level': [f"{mid:.2f}" for mid in mids],
                'TPO Count': tpo_profile.values,
                'Percentage (%)': percentages
            }).sort_values('TPO Count', ascending=False).reset_index(drop=True)
            tpo_df['Cum_TPO'] = tpo_df['TPO Count'].cumsum()
            tpo_df['Cum_Pct'] = (tpo_df['Cum_TPO'] / total_tpo * 100).round(2) if total_tpo > 0 else 0
            price_df['Cum_TPO_Bottom_Up'] = price_df['TPO'].cumsum()
            price_df['Cum_Pct_Bottom_Up'] = (price_df['Cum_TPO_Bottom_Up'] / total_tpo * 100).round(2) if total_tpo > 0 else 0
            high_to_low = price_df.sort_values('Price Level', ascending=False)
            high_to_low['Cum_TPO_Top_Down'] = high_to_low['TPO'].cumsum()
            high_to_low['Cum_Pct_Top_Down'] = (high_to_low['Cum_TPO_Top_Down'] / total_tpo * 100).round(2) if total_tpo > 0 else 0
            cum_df = price_df.merge(high_to_low[['Price Level', 'Cum_TPO_Top_Down', 'Cum_Pct_Top_Down']], on='Price Level', suffixes=('_bottom', '_top'))
            prices = price_df['Price Level'].tolist()
            if len(prices) == 0:
                high_tpo_pct = 0.0
                low_tpo_pct = 100.0
                imbalance = 0.0
                supply_check_status = "No Data"
            else:
                price_threshold_high = np.percentile(prices, 70)
                high_tpo = sum(tpo for p, tpo in zip(prices, price_df['TPO'].tolist()) if p >= price_threshold_high)
                low_tpo = total_tpo - high_tpo
                high_tpo_pct = (high_tpo / total_tpo * 100) if total_tpo > 0 else 0
                low_tpo_pct = 100 - high_tpo_pct
                supply_check_status = "Heavy TPO Supply" if high_tpo_pct > 35 else "Balanced TPO" if high_tpo_pct > 25 else "Demand Heavy TPO"
                imbalance = high_tpo_pct - low_tpo_pct
            key = f"{start.date()} to {end.date()}"
            tpo_dfs[key] = tpo_df
            pocs[key] = poc_price
            total_tpos[key] = total_tpo
            va_highs[key] = va_high
            va_lows[key] = va_low
            va_diffs[key] = va_diff
            va_tpo_pcts[key] = va_tpo_pct
            cum_dfs[key] = cum_df
            supply_checks[key] = {
                'high_tpo_pct': round(high_tpo_pct, 1),
                'low_tpo_pct': round(low_tpo_pct, 1),
                'imbalance': round(imbalance, 1),
                'supply_check': supply_check_status
            }
            prices_plot = mids
            tpos = tpo_profile.values
            fig.add_trace(go.Bar(x=prices_plot, y=tpos, name=key))
        fig.update_layout(
            title='TPO Profile (Market Profile): Price Levels vs. TPO Count',
            xaxis_title='Price Level',
            yaxis_title='TPO Count (Time at Price)',
            bargap=0.1,
            showlegend=True,
            height=500
        )
        main_key = list(tpo_dfs.keys())[0] if tpo_dfs else None
        if main_key:
            main_tpo_df = tpo_dfs.get(main_key, pd.DataFrame())
            main_cum_df = cum_dfs.get(main_key, pd.DataFrame())
            main_tpoc = pocs.get(main_key, 'N/A')
            main_total_tpo = total_tpos.get(main_key, 0)
            main_supply = supply_checks.get(main_key, {})
            main_va_high = va_highs.get(main_key, 'N/A')
            main_va_low = va_lows.get(main_key, 'N/A')
            main_va_diff = va_diffs.get(main_key, 0)
            main_va_tpo_pct = va_tpo_pcts.get(main_key, 0)
            return (fig, {}, [], main_tpo_df, main_tpoc, main_total_tpo,
                    main_cum_df, main_supply, main_va_high, main_va_low, main_va_diff, main_va_tpo_pct)
        else:
            empty_df = pd.DataFrame(columns=['Price Level', 'TPO Count', 'Percentage (%)'])
            return (fig, {}, [], empty_df, 'N/A', 0, empty_df,
                    {'high_tpo_pct': 0.0, 'low_tpo_pct': 100.0, 'imbalance': 0.0, 'supply_check': 'No Data'},
                    'N/A', 'N/A', 0, 0)

    def analyze_ta(self, stock_symbol):
        ta_summary = []
        rsi_fig = make_subplots(rows=1, cols=1)
        macd_fig = make_subplots(rows=1, cols=1)
        bb_fig = make_subplots(rows=1, cols=1)
        stoch_fig = make_subplots(rows=1, cols=1)
        adx_fig = make_subplots(rows=1, cols=1)
        vwap_fig = make_subplots(rows=1, cols=1)
        def colour(signal):
            return 'danger' if signal in ('Overbought', 'Oversold', 'Bearish') else 'success'
        try:
            if self.df.empty or 'Close' not in self.df.columns:
                return ta_summary, [rsi_fig, macd_fig, bb_fig, stoch_fig, adx_fig, vwap_fig]
            if 'RSI' in self.df.columns:
                rsi_clean = self.df['RSI'].dropna()
                if not rsi_clean.empty:
                    rsi_fig.add_trace(
                        go.Scatter(x=self.df['Date'], y=self.df['RSI'],
                                name='RSI', line=dict(color='royalblue'))
                    )
                    rsi_fig.update_layout(title=f'RSI for {stock_symbol}', height=400)
                    rsi_fig.add_hline(y=70, line_dash='dash', line_color='red', annotation_text='Overbought')
                    rsi_fig.add_hline(y=30, line_dash='dash', line_color='green', annotation_text='Oversold')
                    rsi_fig.add_hline(y=50, line_dash='dot', line_color='grey', annotation_text='Neutral')
                    latest_rsi = float(rsi_clean.iloc[-1])
                    rsi_signal = ('Overbought' if latest_rsi > 70 else
                                'Oversold' if latest_rsi < 30 else 'Neutral')
                    ta_summary.append({
                        'title': 'RSI Analysis',
                        'text': [f'Latest RSI: {latest_rsi:.2f}', f'Signal: {rsi_signal}'],
                        'badge': rsi_signal,
                        'color': colour(rsi_signal)
                    })
                    rsi_fig.add_annotation(
                        x=self.df['Date'].iloc[-1], y=latest_rsi,
                        text=f'{latest_rsi:.1f} ({rsi_signal})',
                        showarrow=True
                    )
            if all(c in self.df.columns for c in ('MACD', 'MACD_Signal')):
                macd_clean = self.df['MACD'].dropna()
                if not macd_clean.empty:
                    macd_fig.add_trace(go.Scatter(x=self.df['Date'], y=self.df['MACD'],
                                                name='MACD', line=dict(color='blue')))
                    macd_fig.add_trace(go.Scatter(x=self.df['Date'], y=self.df['MACD_Signal'],
                                                name='Signal', line=dict(color='red')))
                    macd_fig.update_layout(title=f'MACD for {stock_symbol}', height=400)
                    macd_fig.add_hline(y=0, line_dash='dot', line_color='grey', annotation_text='Zero')
                    latest_macd = float(self.df['MACD'].iloc[-1])
                    latest_signal = float(self.df['MACD_Signal'].iloc[-1])
                    macd_signal = 'Bullish' if latest_macd > latest_signal else 'Bearish'
                    ta_summary.append({
                        'title': 'MACD Analysis',
                        'text': [f'Momentum: {macd_signal}', f'MACD: {latest_macd:.4f} | Signal: {latest_signal:.4f}'],
                        'badge': macd_signal,
                        'color': colour(macd_signal)
                    })
                    macd_fig.add_annotation(
                        x=self.df['Date'].iloc[-1], y=latest_macd,
                        text=f'{latest_macd:.3f} ({macd_signal})',
                        showarrow=True
                    )
            if all(c in self.df.columns for c in ('BB_Upper', 'BB_Lower', 'BB_Mid')):
                bb_clean = self.df['BB_Upper'].dropna()
                if not bb_clean.empty:
                    for col, name, clr in zip(
                            ['BB_Upper', 'BB_Mid', 'BB_Lower'],
                            ['Upper', 'Middle', 'Lower'],
                            ['red', 'blue', 'green']):
                        bb_fig.add_trace(go.Scatter(
                            x=self.df['Date'], y=self.df[col],
                            name=f'BB {name}', line=dict(color=clr)))
                    bb_fig.add_trace(go.Scatter(
                        x=self.df['Date'], y=self.df['Close'],
                        name='Close', line=dict(color='black')))
                    bb_fig.update_layout(title=f'Bollinger Bands for {stock_symbol}', height=400)
                    latest_close = float(self.df['Close'].iloc[-1])
                    latest_up = float(self.df['BB_Upper'].iloc[-1])
                    latest_low = float(self.df['BB_Lower'].iloc[-1])
                    bb_signal = ('Above Upper (Overbought)' if latest_close > latest_up else
                                'Below Lower (Oversold)' if latest_close < latest_low else
                                'Within Bands (Range-bound)')
                    ta_summary.append({
                        'title': 'Bollinger Bands',
                        'text': [f'Position: {bb_signal}', f'Close: {latest_close:.2f} | Upper: {latest_up:.2f} | Lower: {latest_low:.2f}'],
                        'badge': bb_signal.split()[0],
                        'color': colour(bb_signal.split()[0])
                    })
                    bb_fig.add_annotation(
                        x=self.df['Date'].iloc[-1], y=latest_close,
                        text=f'Close: ₹{latest_close:.2f} ({bb_signal})',
                        showarrow=True
                    )
            if all(c in self.df.columns for c in ('Stoch_K', 'Stoch_D')):
                stoch_clean = self.df['Stoch_K'].dropna()
                if not stoch_clean.empty:
                    stoch_fig.add_trace(go.Scatter(x=self.df['Date'], y=self.df['Stoch_K'],
                                                name='%K', line=dict(color='blue')))
                    stoch_fig.add_trace(go.Scatter(x=self.df['Date'], y=self.df['Stoch_D'],
                                                name='%D', line=dict(color='red')))
                    stoch_fig.update_layout(title=f'Stochastic Oscillator for {stock_symbol}', height=400)
                    stoch_fig.add_hline(y=80, line_dash='dash', line_color='red', annotation_text='Overbought')
                    stoch_fig.add_hline(y=20, line_dash='dash', line_color='green', annotation_text='Oversold')
                    latest_k = float(self.df['Stoch_K'].iloc[-1])
                    latest_d = float(self.df['Stoch_D'].iloc[-1])
                    stoch_signal = ('Overbought' if latest_k > 80 and latest_k > latest_d else
                                    'Oversold' if latest_k < 20 and latest_k < latest_d else 'Neutral')
                    ta_summary.append({
                        'title': 'Stochastic Oscillator',
                        'text': [f'Signal: {stoch_signal}', f'%K: {latest_k:.1f} | %D: {latest_d:.1f}'],
                        'badge': stoch_signal,
                        'color': colour(stoch_signal)
                    })
                    stoch_fig.add_annotation(
                        x=self.df['Date'].iloc[-1], y=latest_k,
                        text=f'%K: {latest_k:.1f} ({stoch_signal})',
                        showarrow=True
                    )
            if all(c in self.df.columns for c in ('ADX', '+DI', '-DI')):
                adx_clean = self.df['ADX'].dropna()
                if not adx_clean.empty:
                    adx_fig.add_trace(go.Scatter(x=self.df['Date'], y=self.df['ADX'],
                                                name='ADX', line=dict(color='blue')))
                    adx_fig.add_trace(go.Scatter(x=self.df['Date'], y=self.df['+DI'],
                                                name='+DI', line=dict(color='green')))
                    adx_fig.add_trace(go.Scatter(x=self.df['Date'], y=self.df['-DI'],
                                                name='-DI', line=dict(color='red')))
                    adx_fig.update_layout(title=f'ADX for {stock_symbol}', height=400)
                    adx_fig.add_hline(y=25, line_dash='dash', line_color='blue', annotation_text='Strong Trend')
                    adx_fig.add_hline(y=20, line_dash='dot', line_color='orange', annotation_text='Weak Trend')
                    latest_adx = float(self.df['ADX'].iloc[-1])
                    adx_signal = ('Strong Trend' if latest_adx > 25 else
                                'Weak Trend' if latest_adx < 20 else 'Developing Trend')
                    ta_summary.append({
                        'title': 'ADX Analysis',
                        'text': [f'Trend Strength: {adx_signal}', f'ADX Value: {latest_adx:.2f}'],
                        'badge': adx_signal.split()[0],
                        'color': 'success' if latest_adx > 25 else 'warning'
                    })
                    adx_fig.add_annotation(
                        x=self.df['Date'].iloc[-1], y=latest_adx,
                        text=f'{latest_adx:.1f} ({adx_signal})',
                        showarrow=True
                    )
            if 'VWAP' in self.df.columns:
                vwap_fig.add_trace(go.Scatter(x=self.df['Date'], y=self.df['Close'],
                                            name='Close', line=dict(color='black')))
                vwap_fig.add_trace(go.Scatter(x=self.df['Date'], y=self.df['VWAP'],
                                            name='VWAP', line=dict(color='blue')))
                vwap_fig.update_layout(title=f'VWAP vs Close for {stock_symbol}', height=400)
                latest_close = float(self.df['Close'].iloc[-1])
                latest_vwap = float(self.df['VWAP'].iloc[-1])
                vwap_signal = 'Above VWAP (Bullish)' if latest_close > latest_vwap else 'Below VWAP (Bearish)'
                ta_summary.append({
                    'title': 'VWAP Analysis',
                    'text': [f'Position: {vwap_signal}', f'Close − VWAP: {latest_close - latest_vwap:+.2f}'],
                    'badge': vwap_signal.split()[0],
                    'color': colour(vwap_signal.split()[0])
                })
                vwap_fig.add_annotation(
                    x=self.df['Date'].iloc[-1], y=latest_close,
                    text=f'Close: ₹{latest_close:.2f} ({vwap_signal})',
                    showarrow=True
                )
        except Exception as e:
            logging.error(f'Error in analyze_ta for {stock_symbol}: {e}', exc_info=True)
        return ta_summary, [rsi_fig, macd_fig, bb_fig, stoch_fig, adx_fig, vwap_fig]

@server.route('/')
def index():
    return send_from_directory('build', 'index.html')

@server.route('/<path:path>')
def static_files(path):
    return send_from_directory('build', path)

@server.route('/stream')
def stream_view():
    symbol = request.args.get('channel', 'RELIANCE.NS')
    q = get_or_create_queue(symbol)
    def event_stream():
        if not q:
            yield f"data: {json.dumps({'error':'Stream not ready'})}\n\n"
            return
        while True:
            try:
                payload = q.get(timeout=35)
                if payload.get('symbol') == symbol:
                    yield f"data: {json.dumps(payload)}\n\n"
                else:
                    yield f"data: {json.dumps({'heartbeat':1})}\n\n"
            except queue.Empty:
                yield f"data: {json.dumps({'heartbeat':1})}\n\n"
    return Response(event_stream(),
                    mimetype="text/event-stream",
                    headers={'Cache-Control':'no-cache',
                             'Access-Control-Allow-Origin':'*'})

@server.route('/analyze', methods=['POST'])
def analyze_api():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        filename = file.filename
        contents = file.read()
        if not contents:
            return jsonify({'error': 'Empty file uploaded'}), 400
        content_string = base64.b64encode(contents).decode('utf-8')
        contents_full = f"data:text/csv;base64,{content_string}"
        params = json.loads(request.form.get('params', '{}') or '{}')
        peak_diff_dates = params.get('peakDiffDates')
        def to_int(val, default):
            try:
                return int(float(val)) if val else default
            except Exception:
                return default
        window_size = to_int(params.get('windowSize'), 10)
        algo = params.get('algorithm', 'kmeans')
        n_clusters = to_int(params.get('clusters'), 3)
        days_ahead = to_int(params.get('daysAhead'), 10)
        rsi_period = to_int(params.get('rsiPeriod'), 14)
        adx_period = to_int(params.get('adxPeriod'), 14)
        df = parse_contents(contents_full, filename)
        if df.empty:
            return jsonify({'error': 'Empty dataframe after parsing'}), 400
        if 'Date' not in df.columns or 'Close' not in df.columns:
            return jsonify({'error': 'Missing Date/Close columns'}), 400
        symbol = extract_symbol(df)
        yf_data = fetch_yfinance_data(symbol)
        market_cap_cr_val = float(yf_data.get('market_cap_cr') or 0.0)
        try:
            outstanding_shares = int(yf_data.get('outstanding_shares')) if yf_data.get('outstanding_shares') else None
        except Exception:
            outstanding_shares = None
        yf_hist = yf_data['historical_data'].reset_index()
        if not yf_hist.empty:
            yf_hist['Date'] = pd.to_datetime(yf_hist['Date']).dt.tz_localize(None)
            merge_cols = ['Open', 'High', 'Low', 'Close', 'Volume']
            for col in merge_cols:
                if col not in df.columns:
                    df[col] = np.nan
            df = df.merge(yf_hist[['Date'] + merge_cols], on='Date', how='left', suffixes=('', '_yf'))
            for col in merge_cols:
                yf_col = f'{col}_yf'
                if yf_col in df.columns:
                    df[col] = df[yf_col].combine_first(df[col])
                    df.drop(columns=[yf_col], inplace=True)
        for col in ['Open', 'High', 'Low']:
            if col not in df.columns or df[col].isna().all():
                df[col] = df['Close']
            else:
                df[col] = df[col].fillna(df['Close'])
        if 'Volume' not in df.columns or df['Volume'].isna().all():
            df['Volume'] = 1
        else:
            df['Volume'] = df['Volume'].fillna(1).astype(int)
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date']).sort_values('Date').reset_index(drop=True)
        for col in ['Open', 'High', 'Low', 'Close', 'Volume', 'Delivery',
                    'Cumulative Future OI', 'Future OI Change %', 'PCR']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        analyser = StockAnalyzer(df, market_cap_cr_val, outstanding_shares, yf_data, symbol)
        analyser.compute_indicators(rsi_period=rsi_period, adx_period=adx_period)
        date_range = [(df['Date'].min(), df['Date'].max())]
        vol_fig, _, _, vp_df, poc, total_vol, cum_vp, sc_vp, vah_vol, val_vol, vad_vol, vap_vol, peak_diff = analyser.compute_volume_profile(
            date_ranges=date_range, peak_diff_dates=peak_diff_dates)
        oi_fig, _, _, oi_df, poi, total_oi, cum_oi, sc_oi, vah_oi, val_oi, vad_oi, vap_oi = analyser.compute_oi_profile(date_ranges=date_range)
        tpo_fig, _, _, tpo_df, tpoc, total_tpo, cum_tpo, sc_tpo, vah_tpo, val_tpo, vad_tpo, vap_tpo = analyser.compute_tpo_profile(date_ranges=date_range)
        analyser.find_qualifying_windows(window_size=window_size)
        analyser.cluster_patterns(algorithm=algo, n_clusters=n_clusters)
        cluster_fig = analyser.plot_clusters(symbol, algo)
        pcr_fig, pcr_stats = analyser.analyze_pcr_trends(symbol)
        ta_summary, ta_figs = analyser.analyze_ta(symbol)
        rsi_fig, macd_fig, bb_fig, stoch_fig, adx_fig, vwap_fig = ta_figs
        wyckoff_fig, wyckoff_data = analyser.analyze_wyckoff(symbol)
        latest_close = float(analyser.df['Close'].iloc[-1])
        window_size = 10
        if len(analyser.df) >= window_size:
            recent_delivery_sum = analyser.df['Delivery_Value'].tail(window_size).sum()
            recent_delivery_sum_cr = recent_delivery_sum / 10000000
            delivery_threshold_cr = 0.015 * market_cap_cr_val
            delivery_satisfied = recent_delivery_sum > delivery_threshold_cr
            delivery_pct_of_mc = (recent_delivery_sum / (market_cap_cr_val * 10000000)) * 100
            delivery_status = f"Delivery: ₹{recent_delivery_sum_cr:.2f} Cr ({delivery_pct_of_mc:.2f}% of MC) – {'Satisfied (>1.5%)' if delivery_satisfied else f'Not met (need >₹{delivery_threshold_cr:.2f} Cr)'}"
            recent_oi_start = analyser.df['Cumulative Future OI'].iloc[-window_size]
            recent_oi_end = analyser.df['Cumulative Future OI'].iloc[-1]
            oi_increase_pct = ((recent_oi_end - recent_oi_start) / recent_oi_start * 100) if recent_oi_start > 0 else 0
            oi_satisfied = oi_increase_pct > 10
            oi_status = f"OI Increase: {oi_increase_pct:.2f}% – {'Satisfied (>10%)' if oi_satisfied else 'Not met (need >10%)'}"
            both_satisfied = delivery_satisfied and oi_satisfied
            message = "Conditions: Satisfied" if both_satisfied else "Conditions: Partial" if delivery_satisfied or oi_satisfied else "Conditions: Not Met"
            color = 'success' if both_satisfied else 'warning' if delivery_satisfied or oi_satisfied else 'danger'
        else:
            message = "Conditions: Insufficient Data"
            color = 'secondary'
            delivery_status = "N/A"
            oi_status = "N/A"
        delivery_check = {
            'message': message,
            'color': color,
            'delivery_status': delivery_status,
            'oi_status': oi_status
        }
        response = {
            'summary': {
                'metrics': {
                    'symbol': symbol,
                    'outstanding_shares': str(outstanding_shares) if outstanding_shares else 'N/A',
                    'current_price': str(yf_data.get('current_price', 'N/A')),
                    'market_cap_cr': f'{market_cap_cr_val:.2f}' if market_cap_cr_val else 'N/A',
                    '52w_high_low': f"{yf_data.get('52w_high', 'N/A')} / {yf_data.get('52w_low', 'N/A')}"
                },
                'delivery_check': delivery_check,
                'outlook': {
                    'date': str(analyser.df['Date'].iloc[-1].date()),
                    'close': latest_close,
                    'ma5': float(analyser.df['MA5'].iloc[-1] or 0),
                    'pcr': float(analyser.df['PCR'].iloc[-1] or 1),
                    'pcr_ma5': float(analyser.df['PCR'].rolling(5).mean().iloc[-1] or 1),
                    'wyckoff_event': str(analyser.df['Wyckoff_Event'].iloc[-1]),
                    'wyckoff_outlook': ('Bullish' if any(x in str(analyser.df['Wyckoff_Event'].iloc[-1]) for x in ['Spring', 'Sign of Strength']) else
                                        'Bearish' if any(x in str(analyser.df['Wyckoff_Event'].iloc[-1]) for x in ['Upthrust', 'Sign of Weakness']) else 'Neutral'),
                    'oi_divergence': (
                        'Bearish (OI up, Price down) – potential bullish turnaround' if analyser.df['Cumulative Future OI'].pct_change().tail(5).mean() > 0.05 and analyser.df['Close'].pct_change().tail(5).mean() < -0.01 else
                        'Bullish (OI down, Price up) – potential bearish turnaround' if analyser.df['Cumulative Future OI'].pct_change().tail(5).mean() < -0.05 and analyser.df['Close'].pct_change().tail(5).mean() > 0.01 else 'N/A')
                },
                'pattern': {
                    'recent_foi': [p[2] for p in analyser.patterns[-10:]] if len(analyser.patterns) >= 10 else [p[2] for p in analyser.patterns],
                    'cluster_match': f'Cluster {int(np.mean(analyser.labels)) if analyser.labels is not None and len(analyser.labels) else 0}',
                    'expected_change': 2.5,
                    'guidance': ('Bullish' if str(analyser.df['Wyckoff_Event'].iloc[-1]) == 'Sign of Strength' else
                                 'Bearish' if str(analyser.df['Wyckoff_Event'].iloc[-1]) == 'Sign of Weakness' else 'Neutral'),
                    'wyckoff_event': str(analyser.df['Wyckoff_Event'].iloc[-1])
                }
            },
            'volume': {
                'data': vp_df.to_dict('records') if not vp_df.empty else [],
                'cumulative_data': cum_vp.to_dict('records') if not cum_vp.empty else [],
                'poc': poc, 'total_vol': int(total_vol),
                'top3_pct': float(vp_df.head(3)['Percentage (%)'].sum()) if not vp_df.empty and 'Percentage (%)' in vp_df.columns else 0,
                'va_high': vah_vol, 'va_low': val_vol, 'va_diff': round(vad_vol, 2),
                'peak_diff': round(peak_diff, 2), 'va_vol_pct': round(vap_vol, 1),
                'supply_check': sc_vp,
                'date_range': f"{date_range[0][0].date()} to {date_range[0][1].date()}",
                'plot': json.loads(vol_fig.to_json()) if vol_fig else {}
            },
            'oi_profile': {
                'data': oi_df.to_dict('records') if not oi_df.empty else [],
                'cumulative_data': cum_oi.to_dict('records') if not cum_oi.empty else [],
                'poi': poi, 'total_oi': int(total_oi),
                'top3_pct': float(oi_df.head(3)['Percentage (%)'].sum()) if not oi_df.empty and 'Percentage (%)' in oi_df.columns else 0,
                'va_high': vah_oi, 'va_low': val_oi, 'va_diff': round(vad_oi, 2), 'va_oi_pct': round(vap_oi, 1),
                'supply_check': sc_oi,
                'date_range': f"{date_range[0][0].date()} to {date_range[0][1].date()}",
                'plot': json.loads(oi_fig.to_json()) if oi_fig else {}
            },
            'tpo_profile': {
                'data': tpo_df.to_dict('records') if not tpo_df.empty else [],
                'cumulative_data': cum_tpo.to_dict('records') if not cum_tpo.empty else [],
                'tpoc': tpoc, 'total_tpo': int(total_tpo),
                'top3_pct': float(tpo_df.head(3)['Percentage (%)'].sum()) if not tpo_df.empty and 'Percentage (%)' in tpo_df.columns else 0,
                'va_high': vah_tpo, 'va_low': val_tpo, 'va_diff': round(vad_tpo, 2), 'va_tpo_pct': round(vap_tpo, 1),
                'supply_check': sc_tpo,
                'date_range': f"{date_range[0][0].date()} to {date_range[0][1].date()}",
                'plot': json.loads(tpo_fig.to_json()) if tpo_fig else {}
            },
            'clustering': {'plot': json.loads(cluster_fig.to_json()) if cluster_fig else {}},
            'trends': {'pcr_stats': pcr_stats, 'plot': json.loads(pcr_fig.to_json()) if pcr_fig else {}},
            'technical': {
                'summary': ta_summary,
                'plots': {
                    'rsi': json.loads(rsi_fig.to_json()) if rsi_fig else {},
                    'macd': json.loads(macd_fig.to_json()) if macd_fig else {},
                    'bb': json.loads(bb_fig.to_json()) if bb_fig else {},
                    'stoch': json.loads(stoch_fig.to_json()) if stoch_fig else {},
                    'adx': json.loads(adx_fig.to_json()) if adx_fig else {},
                    'vwap': json.loads(vwap_fig.to_json()) if vwap_fig else {}
                }
            },
            'wyckoff': {
                'overview': wyckoff_data['overview'],
                'recent': wyckoff_data['recent'],
                'plot': json.loads(wyckoff_fig.to_json()) if wyckoff_fig else {}
            },
            'periods': {'data': analyser.qualifying.to_dict('records') if not analyser.qualifying.empty else []},
            'stream_url': f'/stream?channel={symbol}',
            'error': None
        }
        edge_diag = analyser._edge_diagonal_oi(oi_df, latest_close)
        cum_oc_df = analyser._cumulative_oi_open_close(oi_df)
        turnaround = analyser._oi_turnaround_point(oi_df, latest_close)
        response['oi_profile'].update({
            'edge_diagonal': edge_diag,
            'cumulative_open_close': cum_oc_df.to_dict('records'),
            'turnaround_point': turnaround
        })
        response = clean_for_json(response)
        return jsonify(response)
    except Exception as e:
        logging.error(f"API Error: {e}")
        logging.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

dash_app.layout = dbc.Container([
    html.H2("Agentic OI & Delivery Analysis — Dash UI", className="mb-4"),
    html.Hr(),
    dbc.Row([
        dbc.Col([
            html.H6("1) Upload your analysis file (CSV or XLSX)", className="mb-2"),
            dcc.Upload(
                id='upload-data',
                children=html.Div(['Drag and Drop or ', html.A('Select Files')], className="text-center"),
                style={
                    'width': '100%', 'height': '60px', 'lineHeight': '60px',
                    'borderWidth': '2px', 'borderStyle': 'dashed', 'borderRadius': '10px',
                    'textAlign': 'center', 'marginBottom': '20px', 'backgroundColor': '#f8f9fa'
                },
                multiple=False
            ),
            html.Div(id='file-info', className="alert alert-info"),
            html.H6("2) Analysis Parameters", className="mt-4 mb-3"),
            dbc.Label("Choose clustering algorithm"),
            dcc.Dropdown(id='algo', options=[
                {'label': 'KMeans', 'value': 'kmeans'}
            ], value='kmeans', className="mb-3"),
            dbc.Label("Days ahead for price change prediction"),
            dcc.Input(id='days-ahead', type='number', value=10, min=1, step=1, className="form-control mb-3"),
            dbc.Label("Window size (days)"),
            dcc.Input(id='window-size', type='number', value=10, min=2, step=1, className="form-control mb-3"),
            dbc.Label("Number of clusters (KMeans)"),
            dcc.Input(id='n-clusters', type='number', value=3, min=2, step=1, className="form-control mb-3"),
            html.H6("Technical Indicators", className="mt-4 mb-3"),
            dbc.Label("RSI Period"),
            dcc.Input(id='rsi-period', type='number', value=14, min=5, step=1, className="form-control mb-3"),
            dbc.Label("ADX Period"),
            dcc.Input(id='adx-period', type='number', value=14, min=5, step=1, className="form-control mb-3"),
            dbc.Button("Run Analysis", id='raw-btn', color='primary', className='w-100 mt-3 mb-4'),
            dcc.Loading(id="loading-1", children=[html.Div(id="loading-output")], type="default"),
            html.Hr(),
            dcc.Download(id="download-results")
        ], width=4, className="border-end"),
        dbc.Col([
            html.H6("Parsed Preview", className="mb-2"),
            html.Div(id='parsed-columns', className="mb-3"),
            dash_table.DataTable(
                id='preview-table',
                page_size=5,
                style_table={'overflowX': 'auto'},
                style_cell={'textAlign': 'left', 'padding': '8px'},
                style_header={'backgroundColor': 'rgb(230, 230, 230)', 'fontWeight': 'bold'}
            ),
            html.Hr(),
            html.H6("Analysis Results", className="mb-3"),
            dbc.Tabs([
                dbc.Tab(label="Summary & Outlook", tab_id="summary", children=[
                    dbc.Row([
                        dbc.Col(id='summary-metrics', width=12),
                        dbc.Col(id='delivery-check', width=12),
                        dbc.Col(id='outlook', width=6),
                        dbc.Col(id='recent-pattern', width=6)
                    ])
                ]),
                dbc.Tab(label="Volume Profile", tab_id="vp", children=[
                    dbc.Row([
                        dbc.Col([
                            dcc.Graph(id='volume-plot', style={'height': '500px'})
                        ], width=12)
                    ]),
                    html.Div(id='vp-text', className="mt-3"),
                    dash_table.DataTable(id='cum-table-vp', page_size=10)
                ]),
                dbc.Tab(label="OI Profile", tab_id="oi", children=[
                    dbc.Row([
                        dbc.Col([
                            dcc.Graph(id='oi-plot', style={'height': '500px'})
                        ], width=12)
                    ]),
                    html.Div(id='oi-text', className="mt-3"),
                    dash_table.DataTable(id='cum-table-oi', page_size=10)
                ]),
                dbc.Tab(label="TPO/Market Profile", tab_id="tpo", children=[
                    dbc.Row([
                        dbc.Col([
                            dcc.Graph(id='tpo-plot', style={'height': '500px'})
                        ], width=12)
                    ]),
                    html.Div(id='tpo-text', className="mt-3"),
                    dash_table.DataTable(id='cum-table-tpo', page_size=10)
                ]),
                dbc.Tab(label="Clustering", tab_id="cluster", children=[
                    dbc.Row([
                        dbc.Col([
                            dcc.Graph(id='cluster-plot', style={'height': '500px'})
                        ], width=12)
                    ])
                ]),
                dbc.Tab(label="PCR Trends", tab_id="pcr", children=[
                    dbc.Row([
                        dbc.Col([
                            html.Div(id='pcr-stats', className="mb-4")
                        ], width=4),
                        dbc.Col([
                            dcc.Graph(id='pcr-plot', style={'height': '500px'})
                        ], width=8)
                    ])
                ]),
                dbc.Tab(label="Technical Analysis", tab_id="ta", children=[
                    html.Div(id='ta-summary'),
                    dbc.Row([
                        dbc.Col([
                            dcc.Graph(id='rsi-plot', style={'height': '400px'})
                        ], width=6),
                        dbc.Col([
                            dcc.Graph(id='macd-plot', style={'height': '400px'})
                        ], width=6)
                    ]),
                    dbc.Row([
                        dbc.Col([
                            dcc.Graph(id='bb-plot', style={'height': '400px'})
                        ], width=6),
                        dbc.Col([
                            dcc.Graph(id='stoch-plot', style={'height': '400px'})
                        ], width=6)
                    ]),
                    dbc.Row([
                        dbc.Col([
                            dcc.Graph(id='adx-plot', style={'height': '400px'})
                        ], width=6),
                        dbc.Col([
                            dcc.Graph(id='vwap-plot', style={'height': '400px'})
                        ], width=6)
                    ])
                ]),
                dbc.Tab(label="Wyckoff Analysis", tab_id="wyckoff", children=[
                    dbc.Row([
                        dbc.Col([
                            html.Div(id='wyckoff-summary', className="mb-4")
                        ], width=4),
                        dbc.Col([
                            dcc.Graph(id='wyckoff-plot', style={'height': '500px'})
                        ], width=8)
                    ])
                ]),
                dbc.Tab(label="Qualifying Periods", tab_id="qual", children=[
                    dbc.Row([
                        dbc.Col([
                            dash_table.DataTable(
                                id='qualifying-table',
                                page_size=10,
                                style_table={'overflowX': 'auto'},
                                style_cell={'textAlign': 'left', 'padding': '8px'},
                                style_header={'backgroundColor': 'rgb(230, 230, 230)', 'fontWeight': 'bold'}
                            )
                        ], width=12)
                    ])
                ])
            ], className="nav-pills")
        ], width=8)
    ])
], fluid=True)

@dash_app.callback(
    Output('file-info', 'children'),
    Output('parsed-columns', 'children'),
    Output('preview-table', 'data'),
    Output('preview-table', 'columns'),
    Input('upload-data', 'contents'),
    State('upload-data', 'filename')
)
def update_file(contents, filename):
    if contents is None:
        return html.Div("No file uploaded"), "", [], []
    try:
        df = parse_contents(contents, filename)
        if df.empty:
            return html.Div("Error: Could not parse file or file is empty"), "", [], []
        cols = df.columns.tolist()
        preview = df.head(5).to_dict('records')
        dt_cols = [{"name": c, "id": c} for c in df.columns]
        info = dbc.Alert([
            html.P(f"Uploaded: {filename}"),
            html.P(f"Rows: {len(df)}, Columns: {len(cols)}"),
            html.P(f"Symbol: {extract_symbol(df)}")
        ], color="info")
        parsed = html.Pre(str(cols), style={'whiteSpace': 'pre-wrap', 'maxHeight': '150px', 'overflowY': 'auto', 'backgroundColor': '#f8f9fa', 'padding': '10px'})
        return info, parsed, preview, dt_cols
    except Exception as e:
        return html.Div(f"Error parsing file: {e}"), "", [], []

if __name__ == '__main__':
    dash_app.run(debug=True, host='0.0.0.0', port=8050)